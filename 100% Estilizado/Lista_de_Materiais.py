import os
import sys
import tkinter as tk
from tkinter import PhotoImage
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell
import pandas as pd
import ctypes

# Lista de materiais atrelados às respostas
materiaisBomba = {
    1: [
        ('1', 'TST96G028', 'PÇ', 'BOMBA ELÉTRICA G3 PRO - 24V - RESERVATÓRIO 2L'),
        ('1', 'TST125910', 'PÇ', 'SUPORTE PARA BOMBA G3'),
        ('6', 'TST011918', 'PÇ', 'PARAFUSO DE FIXAÇÃO DA BOMBA M8 X 1,25 X 35MM'),
        ('6', 'TST208033', 'PÇ', 'PORCA M8'),
        ('12', 'TST208034', 'PÇ', 'ARRUELA M8'),
        ('1', 'TST115122', 'PÇ', 'VÁLVULA DE ALÍVIO 4000 PSI'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST556402', 'PÇ', 'ADAPTADOR 1/4"NPT(M)X 1/8"NPT(F)'),
        ('6', 'TST038000', 'PÇ', 'ARRUELA DE PRESSÃO 3/8'),
        ('1', 'TST17D688', 'PÇ', 'PORTA FUSÍVEL'),
        ('1', '55-2205-04N', 'PÇ', 'FUSIVEL 4A-TIPO LAMINA'),
    ],
    2: [
        ('1', 'TST96G137', 'PÇ', 'BOMBA ELÉTRICA G3 PRO - 24V - RESERVATÓRIO 4L'),
        ('1', 'TST125910', 'PÇ', 'SUPORTE PARA BOMBA G3'),
        ('6', 'TST011918', 'PÇ', 'PARAFUSO DE FIXAÇÃO DA BOMBA M8 X 1,25 X 35MM'),
        ('6', 'TST208033', 'PÇ', 'PORCA M8'),
        ('12', 'TST208034', 'PÇ', 'ARRUELA M8'),
        ('1', 'TST115122', 'PÇ', 'VÁLVULA DE ALÍVIO 4000 PSI'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST556402', 'PÇ', 'ADAPTADOR 1/4"NPT(M)X 1/8"NPT(F)'),
        ('6', 'TST038000', 'PÇ', 'ARRUELA DE PRESSÃO 3/8'),
        ('1', 'TST17D688', 'PÇ', 'PORTA FUSÍVEL'),
        ('1', '55-2205-04N', 'PÇ', 'FUSIVEL 4A-TIPO LAMINA'),
    ],
    3: [
        ('1', 'TST96G138', 'PÇ', 'BOMBA ELÉTRICA G3 PRO - 24V - RESERVATÓRIO 8L'),
        ('1', 'TST125910', 'PÇ', 'SUPORTE PARA BOMBA G3'),
        ('6', 'TST011918', 'PÇ', 'PARAFUSO DE FIXAÇÃO DA BOMBA M8 X 1,25 X 35MM'),
        ('6', 'TST208033', 'PÇ', 'PORCA M8'),
        ('12', 'TST208034', 'PÇ', 'ARRUELA M8'),
        ('1', 'TST115122', 'PÇ', 'VÁLVULA DE ALÍVIO 4000 PSI'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST556402', 'PÇ', 'ADAPTADOR 1/4"NPT(M)X 1/8"NPT(F)'),
        ('6', 'TST038000', 'PÇ', 'ARRUELA DE PRESSÃO 3/8'),
        ('1', 'TST17D688', 'PÇ', 'PORTA FUSÍVEL'),
        ('1', '55-2205-04N', 'PÇ', 'FUSIVEL 4A-TIPO LAMINA'),
    ],
    4: [
        ('1', 'TST25R801', 'PÇ', 'BOMBA G-MINI 24V 1L CONTROLLER'),
        ('1', 'TST125910', 'PÇ', 'SUPORTE PARA BOMBA G3'),
        ('6', 'TST011918', 'PÇ', 'PARAFUSO DE FIXAÇÃO DA BOMBA M8 X 1,25 X 35MM'),
        ('6', 'TST208033', 'PÇ', 'PORCA M8'),
        ('12', 'TST208034', 'PÇ', 'ARRUELA M8'),
        ('1', 'TST115122', 'PÇ', 'VÁLVULA DE ALÍVIO 4000 PSI'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST556402', 'PÇ', 'ADAPTADOR 1/4"NPT(M)X 1/8"NPT(F)'),
        ('6', 'TST038000', 'PÇ', 'ARRUELA DE PRESSÃO 3/8'),
        ('1', 'TST17D688', 'PÇ', 'PORTA FUSÍVEL'),
        ('1', '55-2205-04N', 'PÇ', 'FUSIVEL 4A-TIPO LAMINA'),
    ],
    5: [
        ('1', 'TST25R803', 'PÇ', 'BOMBA G-MINI 12V 1L CONTROLLER'),
        ('1', 'TST125910', 'PÇ', 'SUPORTE PARA BOMBA G3'),
        ('6', 'TST011918', 'PÇ', 'PARAFUSO DE FIXAÇÃO DA BOMBA M8 X 1,25 X 35MM'),
        ('6', 'TST208033', 'PÇ', 'PORCA M8'),
        ('12', 'TST208034', 'PÇ', 'ARRUELA M8'),
        ('1', 'TST115122', 'PÇ', 'VÁLVULA DE ALÍVIO 4000 PSI'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST556402', 'PÇ', 'ADAPTADOR 1/4"NPT(M)X 1/8"NPT(F)'),
        ('6', 'TST038000', 'PÇ', 'ARRUELA DE PRESSÃO 3/8'),
        ('1', 'TST17D688', 'PÇ', 'PORTA FUSÍVEL'),
        ('1', '55-2205-04N', 'PÇ', 'FUSIVEL 4A-TIPO LAMINA'),
    ]
    # Continue preenchendo para os demais casos
}

materiaisDistribuidores = {
    1: [
        ('1', 'TST24Z486', 'PÇ', 'DISTRIBUIDOR  CSP-6(SAÍDAS) SEM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    2: [
        ('1', 'TST24Z487', 'PÇ', 'DISTRIBUIDOR  CSP-8(SAÍDAS) SEM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    3: [
        ('1', 'TST24Z488', 'PÇ', 'DISTRIBUIDOR  CSP-10(SAÍDAS) SEM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    4: [
        ('1', 'TST24Z489', 'PÇ', 'DISTRIBUIDOR  CSP-12(SAÍDAS) SEM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    5: [
        ('1', 'TST24Z490', 'PÇ', 'DISTRIBUIDOR  CSP-14(SAÍDAS) SEM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    6: [
        ('1', 'TST24Z491', 'PÇ', 'DISTRIBUIDOR  CSP-16(SAÍDAS) SEM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    7: [
        ('1', 'TST24Z504', 'PÇ', 'DISTRIBUIDOR  CSP-6(SAÍDAS) COM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    8: [
        ('1', 'TST24Z505', 'PÇ', 'DISTRIBUIDOR  CSP-8(SAÍDAS) COM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    9: [
        ('1', 'TST24Z506', 'PÇ', 'DISTRIBUIDOR  CSP-10(SAÍDAS) COM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    10: [
        ('1', 'TST24Z507', 'PÇ', 'DISTRIBUIDOR  CSP-12(SAÍDAS) COM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    11: [
        ('1', 'TST24Z508', 'PÇ', 'DISTRIBUIDOR  CSP-14(SAÍDAS) COM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    12: [
        ('1', 'TST24Z509', 'PÇ', 'DISTRIBUIDOR  CSP-16(SAÍDAS) COM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    13: [
        ('1', 'TST24Z510', 'PÇ', 'DISTRIBUIDOR  CSP-18(SAÍDAS) COM PINO INDICADOR'),
        ('1', 'TST26A479', 'PÇ', 'SUPORTE PARA DISTRIBUIDOR CSP'),
        ('2', '55-7006-26', 'PÇ', 'PARAFUSO M6S 6x25'),
        ('2', '55-7101-06', 'PÇ', 'PORCA INOX M6'),
        ('4', '55-7150-06', 'PÇ', 'ARRUELA INOX M6'),
        ('1', '55-7012-30', 'PÇ', 'PARAFUSO SEXTAVADO M6S 12X30 - FZB'),
        ('1', '55-7101-12', 'PÇ', 'PORCA DE AÇO M12'),
        ('2', '55-7150-12', 'PÇ', 'ARRUELA M12 - PARTE DO 55-7095-10'),
        ('1', 'TST556419', 'PÇ', 'ADAPTADOR TE 1/8"(M) X (2)-1/8(F)-NPT-LATÃO'),
        ('1', 'TST555888', 'PÇ', 'PINO GRAXEIRO 5000 PSI 1/8 NPT	'),
        ('1', 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
        ('1', 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
    ],
    # Continue preenchendo para os demais casos
}

# Dicionário de saídas para cada distribuidor
saidasDistribuidores = {
    1: 6,
    2: 8,
    3: 10,
    4: 12,
    5: 14,
    6: 16,
    7: 6,
    8: 8,
    9: 10,
    10: 12,
    11: 14,
    12: 16,
    13: 18
}

# Função para somar quantidades e evitar duplicatas
def adicionar_materiais(lista, novos_materiais):
    for novo_item in novos_materiais:
        encontrado = False
        for item in lista:
            if item[1] == novo_item[1]:  # Verifica se o PN (parte numérica) é o mesmo
                nova_qtd = str(int(item[0]) + int(novo_item[0]))  # Soma as quantidades
                lista[lista.index(item)] = (nova_qtd,) + item[1:]  # Atualiza a lista com a nova quantidade
                encontrado = True
                break
        if not encontrado:
            lista.append(novo_item)  # Se não encontrar duplicata, adiciona o novo item

# Função para exportar para Excel com melhor formatação
def exportar_excel(nome_cliente, equipamento, tipo_sistema, tempo_on, tempo_off, resposta_bomba, distribuidores_selecionados, lista_dinamica):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Materiais"

    # Cabeçalho formatado
    ws.merge_cells('A1:E1')
    ws['A1'] = "Lista de Materiais"
    ws['A1'].alignment = Alignment(horizontal="center")
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")  # Cor do texto branca

    # Cor de fundo para algo mais profissional
    ws['A1'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Informações do cliente e equipamento
    campos = ['Nome do Cliente:', 'Equipamento:', 'Tipo de Sistema:', 'Tempo ON:', 'Tempo OFF:']
    valores = [nome_cliente, equipamento, tipo_sistema, tempo_on, tempo_off]
    
    for i, (campo, valor) in enumerate(zip(campos, valores), start=2):
        ws[f'A{i}'] = campo
        ws[f'A{i}'].font = Font(bold=True)  # Negrito
        ws[f'A{i}'].alignment = Alignment(horizontal="center")
        ws[f'B{i}'] = valor
        ws[f'B{i}'].alignment = Alignment(horizontal="center")

    # Combina materiais da bomba e dos distribuidores
    lista_final = []
    adicionar_materiais(lista_final, materiaisBomba[resposta_bomba])
    
    for distribuidor in distribuidores_selecionados:
        if distribuidor in materiaisDistribuidores:
            adicionar_materiais(lista_final, materiaisDistribuidores[distribuidor])
        else:
            print(f"Distribuidor {distribuidor} não encontrado.")

    # Adicionar os materiais da lista dinâmica
    adicionar_materiais(lista_final, lista_dinamica)

    # Cabeçalho da tabela de produtos
    colunas = ['ITEM', 'QTD', 'PN', 'UNID', 'DESCRIÇÃO']
    ws.append(colunas)
    
    # Formatação do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Azul escuro
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = Border(bottom=Side(border_style="thin"))

    for cell in ws[7]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # Adicionar dados e formatar as linhas da lista de materiais
    for i, item in enumerate(lista_final, start=1):
        ws.append([i] + list(item))
        for cell in ws[i+6]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # Borda preta em volta de tudo
    
    # Definir borda preta para todas as células
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # Aplicar formatação em todas as células, incluindo a última linha
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            # Centralizar o conteúdo de todas as células
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Aplicar bordas pretas em torno de todas as células
            cell.border = thin_border
    
    # Ajuste automático da largura das colunas
    for col in ws.columns:
        max_length = 0
        column = None
        for cell in col:
            if not isinstance(cell, MergedCell):  # Ignora células mescladas
                if column is None:
                    column = cell.column_letter  # Define a letra da coluna apenas se a célula não for mesclada
                try:
                    if cell.value:  # Verifica se a célula tem valor
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        if column:  # Ajusta a largura da coluna, se a coluna foi definida
            adjusted_width = max_length + 2  # Margem extra
            ws.column_dimensions[column].width = adjusted_width

    # Gerar o nome do arquivo com base no nome do cliente
    nome_arquivo = f"Lista_de_Materiais_{nome_cliente.replace(' ', '_')}.xlsx"
    
    # Salvar o arquivo
    wb.save(nome_arquivo)
    messagebox.showinfo("Sucesso", f"Arquivo Excel exportado com sucesso!\nNome do arquivo: {nome_arquivo}")




    # Função para obter o caminho correto do arquivo (ícone)
def resource_path(relative_path):
    """Retorna o caminho correto para o arquivo ao usar o PyInstaller"""
    try:
        # Quando rodar o executável, usa o caminho temporário _MEIPASS
        base_path = sys._MEIPASS
    except AttributeError:
        # Quando rodar o script normalmente, usa o caminho original
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)




# Interface gráfica
def criar_janela():
    # Função para criar lista dinâmica com base na quantidade usada
    def gerar_lista_dinamica(quantidade_usada, quantidade_nao_usada, total_saidas_selecionados):
        lista_dinamica = [
            ((quantidade_usada + 1) * 5, 'TST17S553', 'MT', 'MANGUEIRA 8,6X2,3 600 BAR 1/8'),
            ((quantidade_usada * 2) + 2, 'TST17L648', 'PÇ', 'PORCA DO TERMINAL'),
            ((quantidade_usada * 2) + 2, 'TST17R566', 'PÇ', 'TERMINAL 90° 6MM'),
            ((quantidade_usada * 2) + 2, 'TST17R565', 'PÇ', 'TERMINAL RETO 6MM'),
            (quantidade_usada, 'TST17T780', 'PÇ', 'UNIÃO 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
            ((quantidade_usada * 2) + 2, 'TST17T781', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/8" NPT(M)'),
            ((quantidade_usada * 2) + 2, 'TST17T783', 'PÇ', 'ADAPTADOR 90° 6MM(COMPRESSÃO) X 1/4" NPT(M)'),
            (quantidade_usada // 2, 'TST557392', 'PÇ', 'ADAPTADOR CURTO 1/8"NPT(M) X 1/8"NPT(F)'),
            (quantidade_usada // 2, 'TST001299', 'PÇ', 'ADAPTADOR CURTO M8(M) X 1/8"NPT(F)'),
            (quantidade_usada // 2, 'TST563178', 'PÇ', 'ADAPTADOR LONGO-50MM-1/8"NPT(M) X 1/8"NPT(F)'),
            (quantidade_usada // 2, 'TST15K783', 'PÇ', 'ADAPTADOR 90° 1/8"NPT(M) X 1/8"NPT(F)'),
            (quantidade_usada // 2, 'TST557395', 'PÇ', 'ADAPTADOR 45° 1/8"NPT(M) X 1/8"NPT(F)'),
            (quantidade_usada // 2, 'TST557393', 'PÇ', 'ADAPTADOR MEDIO-36MM-1/8"NPT(M) X 1/8"NPT(F)'),
            (quantidade_usada // 2, 'TST150287', 'PÇ', 'ADAPTADOR 1/4 NPT(M) X 3/8 NPT(F)'),
            # PARTE DA MANGUEIRA tava no distribuidor
            (quantidade_nao_usada, 'TST3010401940', 'PÇ', 'BUJÃO DE FECHAMENTO-LUBMANN0'),
            (quantidade_nao_usada, 'TST3010401930', 'PÇ', 'ANEL COBRE BUJÃO-LUBMANN'),
            (quantidade_usada, 'TST17L550', 'PÇ', 'CHECK VALVE DE SAÍDA CSP'),
        ]
        return lista_dinamica

    def enviar_respostas():
        nome_cliente = entry_nome.get()
        equipamento = entry_equipamento.get()
        tipo_sistema = entry_sistema.get()
        tempo_on = entry_tempo_on.get()
        tempo_off = entry_tempo_off.get()
        resposta_bomba = int(var_bomba.get())

        # Distribuidores selecionados com quantidades
        distribuidores_selecionados = []
        for distribuidor, spinbox in distribuidores_spinboxes.items():
            quantidade = int(spinbox.get())
            if quantidade > 0:
                distribuidores_selecionados.extend([distribuidor] * quantidade)  # Adiciona o distribuidor na quantidade especificada

        # Quantidade de saídas usadas
        quantidade_usada = int(entry_quantidade_usada.get())

        # Calcular a quantidade total de saídas dos distribuidores selecionados
        total_saidas_selecionados = sum(saidasDistribuidores[distribuidor] for distribuidor in distribuidores_selecionados)
        quantidade_nao_usada = total_saidas_selecionados - quantidade_usada

        # Exibir a quantidade não usada
        messagebox.showinfo("Quantidade Não Usada", f"Quantidade de saídas que não vão ser usadas: {quantidade_nao_usada}")
        # Exibir a quantidade usada
        messagebox.showinfo("Quantidade Usada", f"Quantidade de saídas que vão ser usadas: {quantidade_usada}")
        # Exibir a quantidade Total de Saídas
        messagebox.showinfo("Total de Saídas", f"Total de Saídas: {total_saidas_selecionados}")

        # Geração da lista de materiais com base na quantidade usada e não usada
        lista_dinamica = gerar_lista_dinamica(quantidade_usada, quantidade_nao_usada, total_saidas_selecionados)

        # Exportar para o Excel
        exportar_excel(nome_cliente, equipamento, tipo_sistema, tempo_on, tempo_off, resposta_bomba, distribuidores_selecionados, lista_dinamica)

    janela = tk.Tk()
    janela.title("Lista de Materiais Graco")

    # Criando estilo para os RadioButtons
    style = ttk.Style()
    style.theme_use('clam')  # Tema que permite personalizações
    style.configure("Custom.TRadiobutton", 
                    background="#2c2c4c",  # Cor de fundo do RadioButton
                    foreground="#ffffff",  # Cor do texto
                    indicatorcolor="#f8c434",  # Cor da bolinha de seleção
                    )

    # Remover efeito hover (passar o mouse por cima)
    style.map("Custom.TRadiobutton",
            background=[('active', '#2c2c4c')],  # Mantém a mesma cor ao passar o mouse
            foreground=[('active', '#ffffff')])  # Mantém a mesma cor do texto

    # Definindo o ID para o aplicativo
    myappid = 'testato.myproduct.subproduct.version'  # String arbitrária
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

    # Adicionar ícone (usando a função resource_path)
    icone_path = resource_path('graco128.ico')
    janela.iconbitmap(icone_path)

    # Estilo da janela
    janela.configure(bg="#2c2c4c")

    # Campos de texto
    tk.Label(janela, text="Nome do Cliente:", bg="#2c2c4c", fg="#ffffff").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    entry_nome = tk.Entry(janela)
    entry_nome.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(janela, text="Equipamento:", bg="#2c2c4c", fg="#ffffff").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    entry_equipamento = tk.Entry(janela)
    entry_equipamento.grid(row=1, column=1, padx=5, pady=5)

    tk.Label(janela, text="Tipo de Sistema:", bg="#2c2c4c", fg="#ffffff").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    entry_sistema = tk.Entry(janela)
    entry_sistema.grid(row=2, column=1, padx=5, pady=5)

    tk.Label(janela, text="Tempo ON:", bg="#2c2c4c", fg="#ffffff").grid(row=3, column=0, padx=5, pady=5, sticky="e")
    entry_tempo_on = tk.Entry(janela)
    entry_tempo_on.grid(row=3, column=1, padx=5, pady=5)

    tk.Label(janela, text="Tempo OFF:", bg="#2c2c4c", fg="#ffffff").grid(row=4, column=0, padx=5, pady=5, sticky="e")
    entry_tempo_off = tk.Entry(janela)
    entry_tempo_off.grid(row=4, column=1, padx=5, pady=5)

    tk.Label(janela, text="Quantidade de saídas usadas:", bg="#2c2c4c", fg="#ffffff").grid(row=5, column=0, padx=5, pady=5, sticky="e")
    entry_quantidade_usada = tk.Entry(janela)
    entry_quantidade_usada.grid(row=5, column=1, padx=5, pady=5)

    # Pergunta sobre a bomba
    tk.Label(janela, text="Selecione a Bomba:", bg="#2c2c4c", fg="#ffffff").grid(row=6, column=0, padx=5, pady=5, sticky="e")
    var_bomba = tk.IntVar()
    bombas = [
        "BOMBA ELÉTRICA G3 PRO - 24V - RESERVATÓRIO 2L",
        "BOMBA ELÉTRICA G3 PRO - 24V - RESERVATÓRIO 4L",
        "BOMBA ELÉTRICA G3 PRO - 24V - RESERVATÓRIO 8L",
        "BOMBA G-MINI 24V 1L CONTROLLER",
        "BOMBA G-MINI 12V 1L CONTROLLER"
    ]

    for i, bomba in enumerate(bombas, start=1):
        ttk.Radiobutton(janela, text=bomba, variable=var_bomba, value=i, style="Custom.TRadiobutton", takefocus=False).grid(row=6+i, column=1, sticky="w")

    # Pergunta sobre os distribuidores
    tk.Label(janela, text="Selecione os Distribuidores:", bg="#2c2c4c", fg="#ffffff").grid(row=12, column=0, padx=5, pady=5, sticky="e")
    distribuidores_spinboxes = {}
    distribuidores_nomes = {
        1: "DISTRIBUIDOR CSP-6(SAÍDAS) SEM PINO INDICADOR",
        2: "DISTRIBUIDOR CSP-8(SAÍDAS) SEM PINO INDICADOR",
        3: "DISTRIBUIDOR CSP-10(SAÍDAS) SEM PINO INDICADOR",
        4: "DISTRIBUIDOR CSP-12(SAÍDAS) SEM PINO INDICADOR",
        5: "DISTRIBUIDOR CSP-14(SAÍDAS) SEM PINO INDICADOR",
        6: "DISTRIBUIDOR CSP-16(SAÍDAS) SEM PINO INDICADOR",
        7: "DISTRIBUIDOR CSP-6(SAÍDAS) COM PINO INDICADOR",
        8: "DISTRIBUIDOR CSP-8(SAÍDAS) COM PINO INDICADOR",
        9: "DISTRIBUIDOR CSP-10(SAÍDAS) COM PINO INDICADOR",
        10: "DISTRIBUIDOR CSP-12(SAÍDAS) COM PINO INDICADOR",
        11: "DISTRIBUIDOR CSP-14(SAÍDAS) COM PINO INDICADOR",
        12: "DISTRIBUIDOR CSP-16(SAÍDAS) COM PINO INDICADOR",
        13: "DISTRIBUIDOR CSP-18(SAÍDAS) COM PINO INDICADOR"
    }

    for i, (num, distribuidor) in enumerate(distribuidores_nomes.items(), start=1):
        tk.Label(janela, text=distribuidor, bg="#2c2c4c", fg="#ffffff").grid(row=12+i, column=1, padx=5, pady=5, sticky="w")
        spinbox = tk.Spinbox(janela, from_=0, to=10, width=5)
        spinbox.grid(row=12+i, column=2, padx=5, pady=5)
        distribuidores_spinboxes[num] = spinbox

    # Botão Enviar
    botao_enviar = tk.Button(janela, text="Enviar", command=enviar_respostas, bg="#f8c434", fg="#2c2c4c", font=("Arial", 10, "bold"))
    botao_enviar.grid(row=27, column=1, pady=10)

    janela.mainloop()

criar_janela()